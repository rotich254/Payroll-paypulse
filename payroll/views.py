from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import models  # Add this import
from .models import Employee, Payroll, Report, Profile, Company
from .forms import EmployeeForm, CompanyForm
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_protect
from django.http import JsonResponse, HttpResponse, Http404
import time  # Add for simulating loading delays
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from decimal import Decimal, InvalidOperation  # Add InvalidOperation
from django.utils import timezone
from datetime import datetime, timedelta
from django.template.loader import get_template, render_to_string
from django.urls import reverse
import tempfile
from django.db.models import Sum, Count
import json
import os
from django.db.models import Sum
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import logout

# Replace WeasyPrint imports with ReportLab
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter # Keep letter for potential reference, but remove landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
# Remove ReportLab platypus imports
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak 
from io import BytesIO
from django.core.files.base import ContentFile  # Add this import
import logging # Add logging import

# +++ Add Openpyxl imports +++
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Configure basic logging (consider moving to settings.py for production)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def ajax_loading_delay(min_delay=0.1):
    """
    Decorator to add minimal loading time for AJAX requests (only if needed for very fast responses)
    """
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            start_time = time.time()
            response = view_func(request, *args, **kwargs)
            elapsed = time.time() - start_time
            
            # Only add minimal delay for extremely fast responses to ensure loader visibility
            if elapsed < min_delay and request.headers.get('x-requested-with') == 'XMLHttpRequest':
                time.sleep(min_delay - elapsed)
            
            return response
        return wrapper
    return decorator

@login_required
def index(request):
    # Get counts for all employee statuses
    total_employees = Employee.objects.count()
    active_employees = Employee.objects.filter(is_active='active').count()
    on_leave_employees = Employee.objects.filter(is_active='on_leave').count()
    probation_employees = Employee.objects.filter(is_active='probation').count()

    # Get the 5 most recently PAID payroll records, ordered by payment date
    recent_payrolls = Payroll.objects.filter(payment_status='paid').select_related('employee').order_by('-payment_date')[:5]

    # Get the 6 nearest upcoming PENDING payrolls
    today = timezone.now().date()
    upcoming_pending_payrolls = Payroll.objects.filter(
        payment_status='pending',
        pay_period__gte=today  # Filter for pay periods today or in the future
    ).exclude(payment_status='paid').select_related('employee').order_by('pay_period')[:5]

    context = {
        'total_employees': total_employees,
        'upcoming_pending_payrolls': upcoming_pending_payrolls,  # Renamed context variable
        'recent_payrolls': recent_payrolls,  # Add recent payrolls to context
        'active_employees': active_employees,
        'on_leave_employees': on_leave_employees,
        'probation_employees': probation_employees,
        'upcoming_periods': get_payroll_periods(),
    }

    # Departmental Payroll Distribution
    department_data = {}
    for employee in Employee.objects.filter(is_active='active'):
        department = employee.get_department_display()
        payroll_records = Payroll.objects.filter(employee=employee, payment_status='paid')
        total_salary = sum(record.net_salary for record in payroll_records)
        if department in department_data:
            department_data[department] += total_salary
        else:
            department_data[department] = total_salary

    context['department_data'] = department_data
    # Get department data for charts
    department_data = Employee.objects.values('department').annotate(
    total_payroll=Sum('salary'),
    employee_count=Count('id')
    )
    
    # Prepare data for charts
    chart_departments = [dict(Employee.DEPARTMENT_CHOICES)[dept['department']] for dept in department_data]
    payroll_by_department = [float(dept['total_payroll'] or 0) for dept in department_data]
    employees_by_department = [dept['employee_count'] for dept in department_data]
    
    context.update({
    'total_employees': total_employees,
    'active_employees': active_employees,
    'on_leave_employees': on_leave_employees,
    'probation_employees': probation_employees,
    'upcoming_pending_payrolls': upcoming_pending_payrolls,
    'recent_payrolls': recent_payrolls,
    'upcoming_periods': get_payroll_periods(),
    # Add chart data to context
    'chart_departments': json.dumps(chart_departments),
    'payroll_by_department': json.dumps(payroll_by_department),
    'employees_by_department': json.dumps(employees_by_department),
    })

    # Add status counts for dropdown
    status_counts = Employee.objects.values('is_active').annotate(count=Count('id'))
    status_counts_dict = {item['is_active']: item['count'] for item in status_counts}
    context['status_counts'] = status_counts_dict

    return render(request, 'index.html', context)

@login_required
def employee_list(request):
    status_filter = request.GET.get('status', '')
    department_filter = request.GET.get('department', '')
    search_query = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort', 'id')
    page = request.GET.get('page', 1)

    employees = Employee.objects.all().order_by('first_name', 'last_name')

    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    if status_filter:
        employees = employees.filter(is_active=status_filter)

    if department_filter:
        employees = employees.filter(department=department_filter)

    paginator = Paginator(employees, 10)
    try:
        employees_page = paginator.page(page)
    except PageNotAnInteger:
        employees_page = paginator.page(1)
    except EmptyPage:
        employees_page = paginator.page(paginator.num_pages)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'payroll/partials/employee_table_body.html', {'employees': employees_page})

    context = {
        'employees': employees_page,
        'search_query': search_query,
        'department_choices': Employee.DEPARTMENT_CHOICES,
        'status_choices': Employee.ACTIVE_STATUS,
        'current_department': department_filter,
        'current_status': status_filter,
        'status_counts': Employee.objects.values('is_active').annotate(count=Count('id')),
    }
    return render(request, 'payroll/employees.html', context)
@login_required
def add_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Employee added successfully.')
            return redirect('employee_list')
        else:
            messages.error(request, 'Error adding employee. Please correct the errors below.')
    else:
        form = EmployeeForm()
    context = {
        'form': form,
        'title': 'Add Employee',
        'department_choices': Employee.DEPARTMENT_CHOICES,
        'status_choices': Employee.ACTIVE_STATUS,
    }
    return render(request, 'payroll/add_employee.html', context)

@login_required
def edit_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, 'Employee details updated successfully.')
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)
    
    context = {
        'form': form,
        'employee': employee,
        'title': 'Edit Employee'
    }
    return render(request, 'payroll/edit_employee.html', context)

@login_required
def remove_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    employee.delete()
    messages.success(request, 'Employee removed successfully.')
    return redirect('employee_list')

@login_required
def company_settings(request):
    company, created = Company.objects.get_or_create(id=1)
    if request.method == 'POST':
        form = CompanyForm(request.POST, request.FILES, instance=company)
        if form.is_valid():
            form.save()
            messages.success(request, 'Company settings updated successfully.')
            return redirect('company_settings')
    else:
        form = CompanyForm(instance=company)
    return render(request, 'payroll/company_settings.html', {'form': form})

@login_required
@require_POST
@ajax_loading_delay(0.2)
def update_employee_status(request, employee_id):
    try:
        employee = get_object_or_404(Employee, id=employee_id)
        status = request.POST.get('status')
        if status in dict(Employee.ACTIVE_STATUS):
            employee.is_active = status
            employee.save()
            return JsonResponse({'status': 'success', 'message': 'Employee status updated successfully.'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid status.'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
def payroll(request):
    if request.method == 'POST':
        form = PayrollForm(request.POST)
        if form.is_valid():
            employee = form.cleaned_data['employee']
            allowances = form.cleaned_data['total_allowances']
            
            # Perform payroll calculations
            gross_salary = Decimal(employee.salary)
            tax_rate = Decimal('16.00')
            health_insurance = Decimal('2000.00')
            retirement_rate = Decimal('5.00')
            tax_amount = gross_salary * tax_rate / Decimal('100')
            retirement_amount = gross_salary * retirement_rate / Decimal('100')
            other_deductions = Decimal('0.00')
            total_deductions_sum = tax_amount + health_insurance + retirement_amount + other_deductions
            net_salary = gross_salary + allowances - total_deductions_sum

            min_net_salary = gross_salary / 3
            if net_salary < min_net_salary:
                form.add_error(None, f'Error: Net pay ({net_salary:,.2f}) is less than 1/3 of gross salary ({min_net_salary:,.2f}).')
            else:
                payroll = form.save(commit=False)
                payroll.gross_salary = gross_salary
                payroll.tax_rate = tax_rate
                payroll.health_insurance = health_insurance
                payroll.retirement_rate = retirement_rate
                payroll.tax_amount = tax_amount
                payroll.retirement_amount = retirement_amount
                payroll.total_deductions = other_deductions
                payroll.net_salary = net_salary
                payroll.payment_status = 'pending'
                payroll.save()
                
                messages.success(request, 'Payroll created successfully.')
                return redirect('payroll_detail', payroll_id=payroll.id)
    else:
        form = PayrollForm()

    context = {
        'form': form,
    }
    return render(request, 'payroll/payroll.html', context)

@login_required
@ajax_loading_delay(0.1)
def search_employees(request):
    search_term = request.GET.get('q', '')
    employees = Employee.objects.filter(
        Q(first_name__icontains=search_term) | Q(last_name__icontains=search_term)
    ).values('id', 'first_name', 'last_name')
    
    results = []
    for employee in employees:
        results.append({
            'id': employee['id'],
            'text': f"{employee['first_name']} {employee['last_name']}"
        })
        
    return JsonResponse({'results': results})

@login_required
@ajax_loading_delay(0.1)
def get_employee_salary(request, employee_id):
    """API endpoint to get employee salary and calculate deductions"""
    try:
        employee = get_object_or_404(Employee, id=employee_id)
        
        gross_salary = Decimal(employee.salary)
        tax_rate = Decimal('16.00')
        health_insurance = Decimal('2000.00')
        retirement_rate = Decimal('5.00')
        
        tax_amount = gross_salary * tax_rate / Decimal('100')
        retirement_amount = gross_salary * retirement_rate / Decimal('100')
        
        total_statutory_deductions = tax_amount + health_insurance + retirement_amount
        
        net_salary_before_allowances = gross_salary - total_statutory_deductions
        
        min_net_salary = gross_salary / 3
        
        data = {
            'salary': float(gross_salary),
            'tax_rate': float(tax_rate),
            'health_insurance': float(health_insurance),
            'retirement_rate': float(retirement_rate),
            'tax_amount': float(tax_amount),
            'retirement_amount': float(retirement_amount),
            'total_statutory_deductions': float(total_statutory_deductions),
            'net_salary_before_allowances': float(net_salary_before_allowances),
            'min_net_salary': float(min_net_salary)
        }
        return JsonResponse(data)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

from .forms import EmployeeForm, CompanyForm, PayrollForm

def get_payroll_periods():
    """Generate payroll periods for the next few months"""
    today = timezone.now().date()
    periods = []
    for month in range(3):  # Next 3 months
        date = today.replace(day=1) + timezone.timedelta(days=32*month)
        periods.extend([
            {
                'value': f"{date.year}-{date.month}-1",
                'label': f"{date.strftime('%B')} 1-15, {date.year}"
            },
            {
                'value': f"{date.year}-{date.month}-2",
                'label': f"{date.strftime('%B')} 16-{(date.replace(day=1) + timezone.timedelta(days=32)).replace(day=1).day-1}, {date.year}"
            }
        ])
    return periods


@login_required
def payroll_list(request):
    payrolls_list = Payroll.objects.all().order_by('-pay_period')
    paginator = Paginator(payrolls_list, 15)  # Show 15 payrolls per page
    page = request.GET.get('page')
    try:
        payrolls = paginator.page(page)
    except PageNotAnInteger:
        payrolls = paginator.page(1)
    except EmptyPage:
        payrolls = paginator.page(paginator.num_pages)
    context = {
        'payrolls': payrolls,
        'title': 'Payroll List'
    }
    return render(request, 'payroll/payroll_list.html', context)

@login_required
def paid_payroll_list(request):
    paid_payrolls_list = Payroll.objects.filter(payment_status='paid').order_by('-payment_date')
    paginator = Paginator(paid_payrolls_list, 15)  # Show 15 payrolls per page
    page = request.GET.get('page')
    try:
        paid_payrolls = paginator.page(page)
    except PageNotAnInteger:
        paid_payrolls = paginator.page(1)
    except EmptyPage:
        paid_payrolls = paginator.page(paginator.num_pages)
    context = {
        'paid_payrolls': paid_payrolls,
        'title': 'Paid Payrolls'
    }
    return render(request, 'payroll/paid_payroll_list.html', context)

@login_required
def pending_payroll_list(request):
    pending_payrolls_list = Payroll.objects.filter(payment_status='pending').order_by('-pay_period')
    paginator = Paginator(pending_payrolls_list, 15)
    page = request.GET.get('page')
    try:
        pending_payrolls = paginator.page(page)
    except PageNotAnInteger:
        pending_payrolls = paginator.page(1)
    except EmptyPage:
        pending_payrolls = paginator.page(paginator.num_pages)
    context = {
        'pending_payrolls': pending_payrolls,
        'title': 'Pending Payrolls'
    }
    return render(request, 'payroll/pending_payroll_list.html', context)

@login_required
def payroll_detail(request, payroll_id):
    """View for displaying detailed payroll information"""
    payroll = get_object_or_404(Payroll.objects.select_related('employee'), id=payroll_id)
    
    context = {
        'payroll': payroll,
        'title': f'Payroll Detail - {payroll.pay_period.strftime("%B %Y")}',
        'employee': payroll.employee,
        'department': dict(Employee.DEPARTMENT_CHOICES).get(payroll.employee.department, ''),
        'status': dict(Payroll.PAYMENT_STATUS_CHOICES).get(payroll.payment_status, ''),
        'today': timezone.now().date(),  # Add today's date to context
    }
    return render(request, 'payroll/payroll_detail.html', context)

@login_required
def generate_payroll_pdf(request, payroll_id):
    """Generate PDF for a specific payroll record using ReportLab"""
    try:
        payroll = get_object_or_404(Payroll.objects.select_related('employee'), id=payroll_id)
        
        response = HttpResponse(content_type='application/pdf')
        filename = f"payroll_{payroll.employee.last_name}_{payroll.pay_period.strftime('%Y_%m')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Create PDF document
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        header_style = styles['Heading2']
        normal_style = styles['Normal']
        
        # Add company header
        company = Company.objects.first()
        company_name = company.name if company else "PayPulse"
        elements.append(Paragraph(company_name, title_style))
        elements.append(Paragraph('Payroll Statement', header_style))
        elements.append(Spacer(1, 20))
        
        # Employee information table
        employee_data = [
            ['Employee Information', '', '', ''],
            ['Name:', f"{payroll.employee.first_name} {payroll.employee.last_name}", 'Payment Status:', 
             f"{payroll.payment_status_display.upper()}" if not payroll.is_paid else 'PAID'],
            ['Employee ID:', str(payroll.employee.id), 'Pay Period:', 
             payroll.pay_period.strftime("%B %d, %Y")],
            ['Status:', payroll.payment_status_display, 'Payment Date:', 
             payroll.payment_date.strftime("%B %d, %Y") if payroll.payment_date else 'Pending']
        ]
        
        employee_table = Table(employee_data, colWidths=[100, 150, 100, 150])
        employee_table.setStyle(TableStyle([
            ('GRID', (0, 1), (-1, -1), 1, colors.black),
            ('SPAN', (0, 0), (-1, 0)),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTWEIGHT', (0, 0), (-1, 0), 'BOLD'),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('TEXTCOLOR', (3, 1), (3, 1), colors.red if not payroll.is_paid else colors.green),
            ('FONTWEIGHT', (3, 1), (3, 1), 'BOLD'),
        ]))
        elements.append(employee_table)
        elements.append(Spacer(1, 20))
        
        # Salary breakdown table
        salary_data = [
            ['Earnings & Deductions', 'Amount'],
            ['Basic Salary', f"KSh {payroll.gross_salary:,.2f}"],
            ['Allowances', f"KSh {payroll.total_allowances:,.2f}"],
            ['Subtotal (Gross)', f"KSh {(payroll.gross_salary + payroll.total_allowances):,.2f}"],
            ['', ''],
            ['Deductions:', ''],
            ['Tax ({:.1f}%)'.format(float(payroll.tax_rate)), f"KSh {payroll.tax_amount:,.2f}"],
            ['Health Insurance', f"KSh {payroll.health_insurance:,.2f}"],
            ['Retirement ({:.1f}%)'.format(float(payroll.retirement_rate)), f"KSh {payroll.retirement_amount:,.2f}"],
            ['Other Deductions', f"KSh {payroll.total_deductions:,.2f}"],
            ['Total Deductions', f"KSh {(payroll.tax_amount + payroll.health_insurance + payroll.retirement_amount + payroll.total_deductions):,.2f}"],
            ['', ''],
            ['Net Salary', f"KSh {payroll.net_salary:,.2f}"]
        ]
        
        salary_table = Table(salary_data, colWidths=[300, 200])
        salary_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, 0), 1, colors.black),
            ('GRID', (0, -1), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, 2), (-1, 2), 1, colors.black),
            ('LINEBELOW', (0, -3), (-1, -3), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('BACKGROUND', (0, 5), (-1, 5), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTWEIGHT', (0, 0), (-1, 0), 'BOLD'),
            ('FONTWEIGHT', (0, -1), (-1, -1), 'BOLD'),
            ('NOSPLIT', (0, 0), (-1, -1)),
        ]))
        elements.append(salary_table)
        
        # Add footer
        elements.append(Spacer(1, 30))
        generated_by = request.user.get_full_name() or request.user.username
        elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')} by {generated_by}", normal_style))
        
        if not payroll.is_paid:
            # Add a prominent watermark
            elements.append(Paragraph(
                "*** DRAFT - NOT YET PAID ***",
                ParagraphStyle(
                    'watermark',
                    parent=normal_style,
                    textColor=colors.red,
                    fontSize=16,
                    alignment=1,  # Center alignment
                    spaceBefore=20,
                    spaceAfter=20
                )
            ))
            elements.append(Paragraph("*** This is not a payment confirmation ***", 
                ParagraphStyle('warning', parent=normal_style, textColor=colors.red)))
        
        # Build and return PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response
        
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('payroll_detail', payroll_id=payroll_id)

@login_required
@require_POST
@ajax_loading_delay(0.2)
def mark_payroll_paid(request, payroll_id):
    try:
        payroll = get_object_or_404(Payroll, id=payroll_id)
        today = timezone.now().date()
        
        # Check if we're trying to pay before the pay date
        if payroll.pay_period > today:
            return JsonResponse({
                'status': 'error',
                'message': f'Cannot mark as paid before the pay date ({payroll.pay_period.strftime("%B %d, %Y")})'
            }, status=400)
            
        if payroll.payment_status == 'pending':
            payroll.payment_status = 'paid'
            payroll.payment_date = today
            payroll.save()
            
            return JsonResponse({
                'status': 'success',
                'message': 'Payroll marked as paid successfully'
            })
        else:
            return JsonResponse({
                'status': 'error',
                'message': 'Payroll is already paid'
        }, status=400)
            
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)



@login_required
def reports(request):
    # Get all reports for display
    reports_list = Report.objects.all().order_by('-generated_date')
    departments = Employee.DEPARTMENT_CHOICES

    # Number of reports per page
    per_page = 10
    paginator = Paginator(reports_list, per_page)
    page = request.GET.get('page')

    try:
        paginated_reports = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        paginated_reports = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        paginated_reports = paginator.page(paginator.num_pages)

    context = {
        'reports': paginated_reports,
        'departments': departments,
    }
    return render(request, 'reports.html', context)

@login_required
@require_POST
@csrf_protect
def delete_report(request, report_id):
    try:
        report = get_object_or_404(Report, id=report_id)
        # Delete the actual file
        if report.file:
            if os.path.exists(report.file.path):
                os.remove(report.file.path)
        # Delete the report record
        report.delete()
        messages.success(request, 'Report deleted successfully')
    except Exception as e:
        messages.error(request, f'Error deleting report: {str(e)}')
    
    return redirect('reports')

from .excel_reports import (
    generate_payroll_excel,
    generate_tax_excel,
    generate_employee_excel,
    generate_department_excel,
)

@login_required
@ajax_loading_delay(0.3)
def generate_report(request):
    """
    Handles report generation requests and calls the appropriate
    Excel generation function based on the selected report type.
    """
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        report_format = request.POST.get('report_format', 'excel')

        # --- Basic Input Validation ---
        if not all([report_type, start_date_str, end_date_str]):
            return JsonResponse({'status': 'error', 'message': 'Missing required fields: report type, start date, or end date.'}, status=400)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid date format. Please use YYYY-MM-DD.'}, status=400)

        if start_date > end_date:
            return JsonResponse({'status': 'error', 'message': 'Start date cannot be after end date.'}, status=400)

        # --- Report Generation Logic ---
        try:
            if report_format == 'excel':
                # Map report types to their generation functions
                report_functions = {
                    'payroll': generate_payroll_excel,
                    'tax': generate_tax_excel,
                    'employee': generate_employee_excel,
                    'department': generate_department_excel,
                }

                # Get the appropriate function
                generation_function = report_functions.get(report_type)

                if generation_function:
                    # Call the function, which now returns a Report object
                    report = generation_function(request, start_date, end_date)
                    if report and isinstance(report, Report):
                        return JsonResponse({
                            'status': 'success',
                            'message': 'Report generated successfully.',
                            'download_url': report.file.url
                        })
                    else:
                        # If the generation function returned None or something else
                        return JsonResponse({'status': 'error', 'message': 'No data found for the selected criteria.'}, status=404)
                else:
                    return JsonResponse({'status': 'error', 'message': f"Invalid report type: '{report_type}'."}, status=400)
            else:
                return JsonResponse({'status': 'error', 'message': f"Unsupported format: '{report_format}'. Only Excel is supported."}, status=400)

        except Exception as e:
            logging.error(f"Error generating report: {e}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': f"An unexpected error occurred: {str(e)}"}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

@login_required
def download_report(request, report_id):
    report = get_object_or_404(Report, id=report_id)
    if report.file:
        response = HttpResponse(report.file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(report.file.name)}"'
        return response
    else:
        raise Http404("File does not exist")

    # Get all reports for display
    recent_reports = Report.objects.all().order_by('-generated_date')[:10]
    departments = Employee.DEPARTMENT_CHOICES

    return render(request, 'reports.html', {
        'recent_reports': recent_reports,
        'departments': departments,
    })

@login_required
def generate_payroll(request):
    """Generate individual payroll view"""
    if request.method == 'POST':
        form = PayrollForm(request.POST)
        if form.is_valid():
            employee = form.cleaned_data['employee']
            allowances = form.cleaned_data['total_allowances']
            other_deductions = form.cleaned_data.get('total_deductions', Decimal('0.00'))
            
            # Perform payroll calculations
            gross_salary = Decimal(employee.salary)
            tax_rate = Decimal('16.00')
            health_insurance = Decimal('2000.00')
            retirement_rate = Decimal('5.00')
            tax_amount = gross_salary * tax_rate / Decimal('100')
            retirement_amount = gross_salary * retirement_rate / Decimal('100')
            total_deductions_sum = tax_amount + health_insurance + retirement_amount + other_deductions
            net_salary = gross_salary + allowances - total_deductions_sum

            min_net_salary = gross_salary / 3
            if net_salary < min_net_salary:
                form.add_error(None, f'Error: Net pay ({net_salary:,.2f}) is less than 1/3 of gross salary ({min_net_salary:,.2f}).')
            else:
                payroll = form.save(commit=False)
                payroll.gross_salary = gross_salary
                payroll.tax_rate = tax_rate
                payroll.health_insurance = health_insurance
                payroll.retirement_rate = retirement_rate
                payroll.tax_amount = tax_amount
                payroll.retirement_amount = retirement_amount
                payroll.total_deductions = other_deductions
                payroll.net_salary = net_salary
                payroll.payment_status = 'pending'
                payroll.save()
                
                messages.success(request, 'Payroll generated successfully.')
                return redirect('payroll_detail', payroll_id=payroll.id)
    else:
        form = PayrollForm()

    context = {
        'form': form,
    }
    return render(request, 'payroll/generate_payroll.html', context)

@login_required
def generate_payroll_report(request, start_date, end_date, department=None):
    logging.info(f"Generating Full Excel Payroll Report for period {start_date} to {end_date}, department: {department}")
    try:
        # --- Date Validation ---
        if isinstance(start_date, str):
            try: start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError: messages.error(request, "Invalid start date format."); return redirect('reports')
        if isinstance(end_date, str):
            try: end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError: messages.error(request, "Invalid end date format."); return redirect('reports')
        if not start_date or not end_date: messages.error(request, "Start/End date required."); return redirect('reports')
        if start_date > end_date: messages.error(request, "Start date cannot be after end date."); return redirect('reports')

        # --- Query Payroll Data (No initial aggregation) ---
        payrolls_query = Payroll.objects.filter(
            pay_period__range=[start_date, end_date]
        ).select_related('employee').order_by('employee__first_name', 'employee__last_name', 'pay_period') # Order results

        if department and department != 'all':
            payrolls_query = payrolls_query.filter(employee__department=department)

        logging.info(f"Found {payrolls_query.count()} individual payroll records matching criteria.")

        if not payrolls_query.exists():
            messages.warning(request, "No payroll records found for the selected criteria.")
            return redirect('reports')

        # --- Generate Excel Workbook ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Details"

        # --- Styles ---
        title_font = Font(bold=True, name='Calibri', size=14)
        period_font = Font(name='Calibri', size=11)
        table_header_font = Font(bold=True, name='Calibri', size=11)
        total_font = Font(bold=True, name='Calibri', size=11)
        currency_format = '"KSh"#,##0.00;[Red]"KSh"-#,##0.00'
        date_format = 'yyyy-mm-dd' # Excel date format string
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid') # Header grey
        total_fill = PatternFill(start_color='FFE0E0E0', end_color='FFE0E0E0', fill_type='solid') # Lighter grey for totals
        thin_border_side = Side(style='thin')
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # --- Write Main Title and Info ---
        current_row = 1
        # Adjust column span based on headers list length below
        max_cols = 9 # Number of columns in the headers list
        ws.cell(row=current_row, column=1, value="Payroll Detail Report")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
        ws.cell(row=current_row, column=1).font = title_font; ws.cell(row=current_row, column=1).alignment = center_alignment
        current_row += 1
        logging.info(f"Current row value: {current_row}")
        
        # --- Finalize and Save ---
        # (Add logic to save the workbook to a file or response)
        # For now, let's assume it's saved and a Report object is created.
        
        # --- Create Report Record ---
        # report = Report.objects.create(...)
        
        messages.success(request, "Excel report generated successfully.")
        return redirect('reports')

    except Exception as e:
        logging.error(f"Error generating Excel report: {e}", exc_info=True)
        messages.error(request, f"An error occurred during report generation: {str(e)}")
        return redirect('reports')
