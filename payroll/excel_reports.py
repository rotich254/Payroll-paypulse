import logging
from datetime import datetime
from io import BytesIO

from django.contrib import messages
from django.db.models import Count, Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from .models import Employee, Payroll, Report
from django.core.files.base import ContentFile

# Configure logging
logging.basicConfig(level=logging.INFO)


def create_styled_workbook(title):
    """Creates a workbook with a styled title and returns the workbook and worksheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Add and style the main title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    return wb, ws


def set_header_styles(ws, headers):
    """Applies styling to header cells and sets column widths."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            bottom=Side(style="thin"),
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
        )
        # Adjust column width
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
    return ws


def generate_payroll_excel(request, start_date, end_date):
    """Generates an Excel report for payroll data and saves it to the database."""
    payrolls = Payroll.objects.filter(
        pay_period__range=[start_date, end_date]
    ).select_related("employee")

    if not payrolls.exists():
        messages.warning(request, "No payroll records found for the selected criteria.")
        return None

    wb, ws = create_styled_workbook("Payroll Detail Report")
    ws.cell(
        row=2,
        column=1,
        value=f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
    ).font = Font(italic=True)

    headers = [
        "Employee ID", "Employee Name", "Department", "Gross Salary",
        "Allowances", "Deductions", "Net Salary", "Pay Period",
    ]
    ws = set_header_styles(ws, headers)

    for row_num, payroll in enumerate(payrolls, 4):
        ws.cell(row=row_num, column=1, value=payroll.employee.id)
        ws.cell(row=row_num, column=2, value=f"{payroll.employee.first_name} {payroll.employee.last_name}")
        ws.cell(row=row_num, column=3, value=payroll.employee.get_department_display())
        ws.cell(row=row_num, column=4, value=payroll.gross_salary)
        ws.cell(row=row_num, column=5, value=payroll.total_allowances)
        ws.cell(row=row_num, column=6, value=payroll.total_deductions)
        ws.cell(row=row_num, column=7, value=payroll.net_salary)
        ws.cell(row=row_num, column=8, value=payroll.pay_period)

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create a Report record
    report_name = f"Payroll_Report_{start_date}_{end_date}.xlsx"
    report = Report.objects.create(
        name=report_name,
        type='payroll',
        period_start=start_date,
        period_end=end_date,
        generated_by=request.user,
    )
    report.file.save(report_name, ContentFile(buffer.read()))

    messages.success(request, "Payroll report generated and saved successfully.")
    
    # Return an HTTP response for download
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={report_name}"
    return response


def generate_tax_excel(request, start_date, end_date):
    """Generates a tax Excel report and saves it to the database."""
    payrolls = Payroll.objects.filter(
        pay_period__range=[start_date, end_date]
    ).select_related("employee")

    if not payrolls.exists():
        messages.warning(request, "No tax records found for the selected criteria.")
        return None

    wb, ws = create_styled_workbook("Tax Report")
    ws.cell(
        row=2,
        column=1,
        value=f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
    ).font = Font(italic=True)

    headers = ["Employee ID", "Employee Name", "Tax Amount", "Pay Period"]
    ws = set_header_styles(ws, headers)

    for row_num, payroll in enumerate(payrolls, 4):
        ws.cell(row=row_num, column=1, value=payroll.employee.id)
        ws.cell(row=row_num, column=2, value=f"{payroll.employee.first_name} {payroll.employee.last_name}")
        ws.cell(row=row_num, column=3, value=payroll.tax_amount)
        ws.cell(row=row_num, column=4, value=payroll.pay_period)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    report_name = f"Tax_Report_{start_date}_{end_date}.xlsx"
    report = Report.objects.create(
        name=report_name,
        type='tax',
        period_start=start_date,
        period_end=end_date,
        generated_by=request.user,
    )
    report.file.save(report_name, ContentFile(buffer.read()))

    messages.success(request, "Tax report generated and saved successfully.")

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={report_name}"
    return response


def generate_employee_excel(request, start_date, end_date):
    """Generates an employee Excel report and saves it to the database."""
    employees = Employee.objects.all()

    if not employees.exists():
        messages.warning(request, "No employees found.")
        return None

    wb, ws = create_styled_workbook("Employee Report")
    headers = [
        "Employee ID", "First Name", "Last Name", "Email",
        "Department", "Hire Date", "Status",
    ]
    ws = set_header_styles(ws, headers)

    for row_num, employee in enumerate(employees, 4):
        ws.cell(row=row_num, column=1, value=employee.id)
        ws.cell(row=row_num, column=2, value=employee.first_name)
        ws.cell(row=row_num, column=3, value=employee.last_name)
        ws.cell(row=row_num, column=4, value=employee.email)
        ws.cell(row=row_num, column=5, value=employee.get_department_display())
        ws.cell(row=row_num, column=6, value=employee.hire_date)
        ws.cell(row=row_num, column=7, value=employee.get_is_active_display())

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    report_name = "Employee_Report.xlsx"
    report = Report.objects.create(
        name=report_name,
        type='employee',
        period_start=start_date,
        period_end=end_date,
        generated_by=request.user,
    )
    report.file.save(report_name, ContentFile(buffer.read()))

    messages.success(request, "Employee report generated and saved successfully.")

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={report_name}"
    return response


def generate_department_excel(request, start_date, end_date):
    """Generates a department summary Excel report and saves it to the database."""
    department_summary = (
        Payroll.objects.filter(pay_period__range=[start_date, end_date])
        .values("employee__department")
        .annotate(
            employee_count=Count("employee", distinct=True),
            total_gross=Sum("gross_salary"),
            total_net=Sum("net_salary"),
        )
    )

    if not department_summary:
        messages.warning(request, "No departmental data found for the selected criteria.")
        return None

    wb, ws = create_styled_workbook("Department Summary Report")
    ws.cell(
        row=2,
        column=1,
        value=f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
    ).font = Font(italic=True)

    headers = [
        "Department", "Employee Count", "Total Gross Salary", "Total Net Salary",
    ]
    ws = set_header_styles(ws, headers)

    for row_num, summary in enumerate(department_summary, 4):
        department_display = dict(Employee.DEPARTMENT_CHOICES).get(
            summary["employee__department"]
        )
        ws.cell(row=row_num, column=1, value=department_display)
        ws.cell(row=row_num, column=2, value=summary["employee_count"])
        ws.cell(row=row_num, column=3, value=summary["total_gross"])
        ws.cell(row=row_num, column=4, value=summary["total_net"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    report_name = f"Department_Summary_{start_date}_{end_date}.xlsx"
    report = Report.objects.create(
        name=report_name,
        type='department',
        period_start=start_date,
        period_end=end_date,
        generated_by=request.user,
    )
    report.file.save(report_name, ContentFile(buffer.read()))

    messages.success(request, "Department summary report generated and saved successfully.")

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={report_name}"
    return response
